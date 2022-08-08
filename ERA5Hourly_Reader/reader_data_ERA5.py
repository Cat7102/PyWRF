# -*- coding: utf-8 -*-
# @Author: Piero
# @Date  : 2022-08-02 18:30:32
# @Desc  : 本脚本用于ERA5 Hourly数据的提取

import netCDF4 as nc
import numpy as np
from openpyxl import Workbook
from datetime import datetime,timedelta

# ERA5数据存储地址
era5_path=r"D:\Data\CFD_Files\CFD_Project_File\ERA5数据\Sh_20210710-20210715_era5_hourly.nc"
# 文件保存路径
xlsx_path=r"D:\Data\CFD_Files\CFD_Project_File\ERA5数据\ERA5_Data.xlsx"

# 主体部分
ncfile=nc.Dataset(era5_path)
# 读取数据
hour_time = ncfile["time"][:]
soil_temp4=ncfile["stl4"][:,6,7]
total_radiation=ncfile["ssrd"][:,6,7]
direct_radiation=ncfile["fdir"][:,6,7]
albedo_IR_diffuse=ncfile["alnid"][:,6,7]
albedo_IR_direct=ncfile["alnip"][:,6,7]
albedo_visible_diffuse=ncfile["aluvd"][:,6,7]
albedo_visible_direct=ncfile["aluvp"][:,6,7]
print(datetime(1900,1,1,0,0,0)+timedelta(hours=int(hour_time[0])+8))
# 创建xlsx
wb=Workbook()
ws = wb["Sheet"]
keylist=["Time","Soil Temperature level 4(K)","Total Solar Radiation(J/m2)","Direct Solar Radiation(J/m2)",
         "Diffuse Infrared Albedo","Direct Infrared Albedo","Diffuse Visible Albedo","Direct Infrared Albedo"]
# 写入首行
for i in range(8):
    ws.cell(1, i+1, keylist[i])
# 写入时间
for i in range(len(hour_time)):
    ws.cell(i+2,1,datetime(1900,1,1,0,0,0)+timedelta(hours=int(hour_time[i])+8))
# 写入其他数据
c=2
for data in [soil_temp4,total_radiation,direct_radiation,albedo_IR_diffuse,albedo_IR_direct,albedo_visible_diffuse,albedo_visible_direct]:
    for i in range(len(data)):
        ws.cell(i+2,c,data[i])
    c+=1
wb.save(xlsx_path)
