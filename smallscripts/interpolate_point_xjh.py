# 将wrf输出值插值到需要的格点上

from lib.Interpolate import interpolate,calculate_wind_direction
import lib.Readtime as Readtime
from wrf import getvar,to_np
from openpyxl import Workbook
import numpy as np
import netCDF4 as nc

#下面修改路径
path=r'D:\Data\WRF-Chem_Files\WRF-Chem_Simulation\LCZ_2021_Jul_HighTemp\wrfout_d03_2021-07-09_00-00-00'

#下面修改站点的纬度，经度，气象站点名字。气象站点名字可以带中文，主要就是为了excel表格的sheet书写
point_list=[(31.2,121.4333,"58367徐家汇")]
ncfile=nc.Dataset(path)
wb=Workbook()
for i in point_list:
    wb.create_sheet(str(i[2]))
    ws = wb[str(i[2])]
    ws.cell(1,1,"时间")
    ws.cell(1,2,"2m温度")
    ws.cell(1,3,"2m湿度")
    ws.cell(1,4,"地表温度")
    ws.cell(1,5,"地表压力")
    ws.cell(1,6,"云层顶部温度")

timelist=Readtime.get_ncfile_time(ncfile)
timestep=3
for i in range(0,Readtime.get_ncfile_alltime(ncfile),timestep):
    t2=to_np(getvar(ncfile,"T2",timeidx=i))-273.15
    rh2=to_np(getvar(ncfile,'rh2',timeidx=i))
    skt=to_np(getvar(ncfile,'TSK',timeidx=i))-273.15
    sp=to_np(getvar(ncfile,'PSFC',timeidx=i))
    ctt=to_np(getvar(ncfile,'ctt',timeidx=i))
    for j in range(len(point_list)):
        float_t2 = interpolate(ncfile, t2, point_list[j][0], point_list[j][1], opt=0)
        float_rh2 = interpolate(ncfile, rh2, point_list[j][0], point_list[j][1], opt=0)
        float_skt = interpolate(ncfile, skt, point_list[j][0], point_list[j][1], opt=0)
        float_sp = interpolate(ncfile, sp, point_list[j][0], point_list[j][1], opt=0)
        float_ctt = interpolate(ncfile, ctt, point_list[j][0], point_list[j][1], opt=0)
        print("当前时间为：{}".format(timelist[i]))
        print("行数为：{}".format(int(i/timestep+2)))
        worksheet=wb[str(point_list[j][2])]
        worksheet.cell(int(i/timestep+2),1,timelist[i])
        worksheet.cell(int(i/timestep+2),2,float_t2)
        worksheet.cell(int(i/timestep+2),3,float_rh2)
        worksheet.cell(int(i/timestep+2),4,float_skt)
        worksheet.cell(int(i/timestep+2),5,float_sp)
        worksheet.cell(int(i/timestep+2),6,float_ctt)
wb.remove(wb['Sheet'])
wb.save("气象信息-徐家汇站点插值.xlsx")




