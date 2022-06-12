# for yrf

import netCDF4 as nc
import openpyxl
from datetime import datetime,timedelta
import numpy as np

workbook=openpyxl.Workbook()
table = workbook["Sheet"]
table.cell(1,1).value="时间"
table.cell(1,2).value="2m温度"
table.cell(1,3).value="10米u"
table.cell(1,4).value="10米v"
table.cell(1,5).value="海温"


data=nc.Dataset(r"D:\era5-yrf.nc")
time=data["time"]
for i in range(len(time)):
    dt = datetime.strptime("1900-01-01 00:00:00", "%Y-%m-%d %H:%M:%S")
    date=dt + timedelta(hours=np.int(time[i]))
    print(date)
    table.cell(i+2,1).value=str(date)
j=2
for v in ["t2m","u10","v10","sst"]:
    f=data[v]
    for i in range(len(time)):
        table.cell(i+2,j).value=f[i,0,0]
    print(j)
    j=j+1
workbook.save("era5-yrf.xlsx")